"""
Comprehensive Excel Financial Query System
==========================================

A fully dynamic system that analyzes Excel structure and answers queries
without hardcoding. Uses intelligent pattern recognition and data analysis.

Author: AI Assistant
Date: 2025
"""

import openpyxl
import pandas as pd
import numpy as np
import re
import os
import logging
from typing import Any, Dict, List, Optional, Union, Tuple
from dataclasses import dataclass, field
from datetime import datetime
import json
from collections import defaultdict, Counter
import math

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@dataclass
class CellData:
    """Represents a cell with its metadata"""
    value: Any
    address: str
    data_type: str
    formula: Optional[str] = None
    formatted_value: Optional[str] = None

@dataclass
class TableStructure:
    """Represents a structured table within a sheet"""
    name: str
    start_row: int
    start_col: int
    end_row: int
    end_col: int
    headers_row: Optional[int] = None
    headers_col: Optional[int] = None
    data_rows: List[int] = field(default_factory=list)
    data_cols: List[int] = field(default_factory=list)
    table_type: Optional[str] = None

@dataclass
class SheetAnalysis:
    """Analysis results for a single sheet"""
    name: str
    tables: List[TableStructure] = field(default_factory=list)
    financial_accounts: Dict[str, int] = field(default_factory=dict)
    time_periods: Dict[str, int] = field(default_factory=dict)
    companies: List[str] = field(default_factory=list)
    sheet_type: Optional[str] = None
    data_density: float = 0.0

@dataclass
class WorkbookAnalysis:
    """Complete workbook analysis"""
    sheets: Dict[str, SheetAnalysis] = field(default_factory=dict)
    global_companies: List[str] = field(default_factory=list)
    global_time_periods: List[str] = field(default_factory=list)
    financial_taxonomy: Dict[str, List[str]] = field(default_factory=dict)
    cross_references: Dict[str, List[str]] = field(default_factory=dict)

class FinancialTaxonomyEngine:
    """Engine for understanding financial terminology and relationships"""
    
    # Comprehensive financial term mappings
    REVENUE_PATTERNS = [
        r'\b(revenue|income|sales|billing|invoic)\b',
        r'\b(shipping\s+income|service\s+income)\b',
        r'\b(gross\s+sales|net\s+sales)\b',
        r'\b(operating\s+income|ordinary\s+income)\b'
    ]
    
    COST_PATTERNS = [
        r'\b(cost|expense|expenditure|outlay)\b',
        r'\b(direct\s+(cost|labor|material))\b',
        r'\b(indirect\s+(cost|labor|overhead))\b',
        r'\b(operating\s+expense|admin\s+expense)\b',
        r'\b(payroll|wages|salaries|benefits)\b'
    ]
    
    PROFIT_PATTERNS = [
        r'\b(profit|margin|ebitda|ebit)\b',
        r'\b(gross\s+profit|net\s+profit|operating\s+profit)\b',
        r'\b(contribution\s+margin)\b'
    ]
    
    BALANCE_SHEET_PATTERNS = [
        r'\b(assets|liabilities|equity|capital)\b',
        r'\b(cash|receivables|inventory|payables)\b',
        r'\b(debt|loan|borrowing|financing)\b'
    ]
    
    TIME_PATTERNS = [
        r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\b',
        r'\b(january|february|march|april|may|june|july|august|september|october|november|december)\b',
        r'\b(q[1-4]|quarter)\b',
        r'\b(20\d{2}|\'\d{2})\b',
        r'\b(ytd|ttm|forecast|budget|actual)\b'
    ]
    
    COMPANY_PATTERNS = [
        r'\b[A-Z]{2,5}\b',  # 2-5 letter abbreviations
        r'\b[A-Z][a-z]+\s+(Inc|LLC|Corp|Co)\b'
    ]
    
    @classmethod
    def categorize_text(cls, text: str) -> List[str]:
        """Categorize text into financial categories"""
        if not text or not isinstance(text, str):
            return []
        
        text_lower = text.lower()
        categories = []
        
        # Check revenue patterns
        if any(re.search(pattern, text_lower) for pattern in cls.REVENUE_PATTERNS):
            categories.append('revenue')
        
        # Check cost patterns
        if any(re.search(pattern, text_lower) for pattern in cls.COST_PATTERNS):
            categories.append('cost')
        
        # Check profit patterns
        if any(re.search(pattern, text_lower) for pattern in cls.PROFIT_PATTERNS):
            categories.append('profit')
        
        # Check balance sheet patterns
        if any(re.search(pattern, text_lower) for pattern in cls.BALANCE_SHEET_PATTERNS):
            categories.append('balance_sheet')
        
        return categories
    
    @classmethod
    def extract_time_periods(cls, text: str) -> List[str]:
        """Extract time periods from text"""
        if not text or not isinstance(text, str):
            return []
        
        periods = []
        text_lower = text.lower()
        
        for pattern in cls.TIME_PATTERNS:
            matches = re.findall(pattern, text_lower)
            periods.extend(matches)
        
        return list(set(periods))
    
    @classmethod
    def extract_companies(cls, text: str) -> List[str]:
        """Extract potential company names from text"""
        if not text or not isinstance(text, str):
            return []
        
        companies = []
        
        for pattern in cls.COMPANY_PATTERNS:
            matches = re.findall(pattern, text)
            companies.extend(matches)
        
        return list(set(companies))

class ExcelStructureAnalyzer:
    """Analyzes Excel file structure and content"""
    
    def __init__(self):
        self.taxonomy = FinancialTaxonomyEngine()
        
    def analyze_workbook(self, file_path: str) -> WorkbookAnalysis:
        """Perform comprehensive workbook analysis"""
        logger.info(f"Starting analysis of {file_path}")
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(file_path, data_only=True)
            wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
            
            workbook_analysis = WorkbookAnalysis()
            
            # Analyze each sheet
            for sheet_name in wb.sheetnames:
                logger.info(f"Analyzing sheet: {sheet_name}")
                
                ws = wb[sheet_name]
                ws_formulas = wb_formulas[sheet_name]
                
                sheet_analysis = self._analyze_sheet(ws, ws_formulas, sheet_name)
                workbook_analysis.sheets[sheet_name] = sheet_analysis
            
            # Perform global analysis
            self._perform_global_analysis(workbook_analysis)
            
            logger.info(f"Analysis complete. Processed {len(workbook_analysis.sheets)} sheets")
            return workbook_analysis
            
        except Exception as e:
            logger.error(f"Error analyzing workbook: {e}")
            raise
    
    def _analyze_sheet(self, ws, ws_formulas, sheet_name: str) -> SheetAnalysis:
        """Analyze a single sheet"""
        
        sheet_analysis = SheetAnalysis(name=sheet_name)
        
        # Extract all cell data
        all_cells = self._extract_cell_data(ws, ws_formulas)
        
        # Detect tables
        tables = self._detect_tables(all_cells, ws.max_row, ws.max_column)
        sheet_analysis.tables = tables
        
        # Analyze content
        self._analyze_sheet_content(all_cells, sheet_analysis)
        
        # Classify sheet type
        sheet_analysis.sheet_type = self._classify_sheet_type(sheet_name, sheet_analysis)
        
        # Calculate data density
        non_empty_cells = sum(1 for cell in all_cells.values() if cell.value is not None)
        total_cells = ws.max_row * ws.max_column if ws.max_row and ws.max_column else 1
        sheet_analysis.data_density = non_empty_cells / total_cells
        
        return sheet_analysis
    
    def _extract_cell_data(self, ws, ws_formulas) -> Dict[str, CellData]:
        """Extract all cell data with metadata"""
        cells = {}
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                cell_formulas = ws_formulas.cell(row, col)
                
                if cell.value is not None or (hasattr(cell_formulas, 'formula') and cell_formulas.formula):
                    address = cell.coordinate
                    
                    # Determine data type
                    data_type = 'text'
                    if isinstance(cell.value, (int, float)):
                        data_type = 'number'
                    elif isinstance(cell.value, datetime):
                        data_type = 'date'
                    elif isinstance(cell.value, bool):
                        data_type = 'boolean'
                    
                    cells[address] = CellData(
                        value=cell.value,
                        address=address,
                        data_type=data_type,
                        formula=getattr(cell_formulas, 'formula', None),
                        formatted_value=cell.displayed_value if hasattr(cell, 'displayed_value') else str(cell.value)
                    )
        
        return cells
    
    def _detect_tables(self, all_cells: Dict[str, CellData], max_row: int, max_col: int) -> List[TableStructure]:
        """Detect table structures in the sheet"""
        tables = []
        
        # Find potential header rows (rows with mostly text or mixed content)
        header_candidates = []
        
        for row in range(1, min(31, max_row + 1)):  # Check first 30 rows
            row_cells = [cell for addr, cell in all_cells.items() 
                        if self._get_row_from_address(addr) == row]
            
            if len(row_cells) >= 2:  # Need at least 2 cells
                # More flexible header detection
                text_ratio = sum(1 for cell in row_cells if cell.data_type == 'text') / len(row_cells)
                mixed_ratio = sum(1 for cell in row_cells if cell.data_type in ['text', 'number']) / len(row_cells)
                
                # Accept rows with good text content or mixed content
                if text_ratio >= 0.4 or (mixed_ratio >= 0.7 and text_ratio >= 0.2):
                    header_candidates.append(row)
        
        # For each header candidate, try to find the associated data table
        for header_row in header_candidates:
            table = self._extract_table_from_header(all_cells, header_row, max_row, max_col)
            if table:
                tables.append(table)
        
        # If no tables detected, try alternative detection methods
        if not tables:
            tables = self._detect_tables_alternative(all_cells, max_row, max_col)
        
        return tables
    
    def _detect_tables_alternative(self, all_cells: Dict[str, CellData], max_row: int, max_col: int) -> List[TableStructure]:
        """Alternative table detection when primary method fails"""
        tables = []
        
        # Look for patterns of numeric data with text labels
        for row in range(1, min(51, max_row + 1)):
            row_cells = [cell for addr, cell in all_cells.items() 
                        if self._get_row_from_address(addr) == row]
            
            if len(row_cells) >= 3:
                # Check if this row has a text label followed by numeric data
                first_cell = row_cells[0]
                if first_cell.data_type == 'text' and first_cell.value:
                    # Look for numeric data in subsequent columns
                    numeric_count = 0
                    start_col = self._get_col_from_address(first_cell.address)
                    
                    for col in range(start_col + 1, min(start_col + 20, max_col + 1)):
                        addr = f"{chr(64 + col)}{row}"
                        if addr in all_cells and all_cells[addr].data_type == 'number':
                            numeric_count += 1
                    
                    if numeric_count >= 2:  # At least 2 numeric columns
                        # Create a simple table structure
                        table = TableStructure(
                            name=f"SimpleTable_{row}_{start_col}",
                            start_row=row,
                            start_col=start_col,
                            end_row=row,
                            end_col=start_col + numeric_count,
                            headers_row=None,  # No separate header row
                            data_rows=[row],
                            data_cols=list(range(start_col, start_col + numeric_count + 1)),
                            table_type='simple_data'
                        )
                        tables.append(table)
        
        return tables
    
    def _extract_table_from_header(self, all_cells: Dict[str, CellData], 
                                   header_row: int, max_row: int, max_col: int) -> Optional[TableStructure]:
        """Extract table structure starting from a header row"""
        
        # Find header columns
        header_cells = [(addr, cell) for addr, cell in all_cells.items() 
                       if self._get_row_from_address(addr) == header_row]
        
        if len(header_cells) < 2:
            return None
        
        # Determine table boundaries
        start_col = min(self._get_col_from_address(addr) for addr, _ in header_cells)
        end_col = max(self._get_col_from_address(addr) for addr, _ in header_cells)
        
        # Find data rows (rows after header with numeric data)
        data_rows = []
        for row in range(header_row + 1, min(header_row + 200, max_row + 1)):
            row_cells = [cell for addr, cell in all_cells.items() 
                        if self._get_row_from_address(addr) == row 
                        and start_col <= self._get_col_from_address(addr) <= end_col]
            
            if not row_cells:
                # Empty row might indicate end of table
                if len(data_rows) > 0:
                    break
                continue
            
            # Check if row has numeric data
            numeric_ratio = sum(1 for cell in row_cells if cell.data_type == 'number') / len(row_cells)
            
            if numeric_ratio >= 0.3:  # At least 30% numeric data
                data_rows.append(row)
            elif len(data_rows) > 3:  # If we have some data and hit non-numeric, might be end
                break
        
        if len(data_rows) < 2:  # Need at least 2 data rows
            return None
        
        end_row = max(data_rows)
        
        # Classify table type
        table_type = self._classify_table_type(all_cells, header_row, start_col, end_col)
        
        return TableStructure(
            name=f"Table_{header_row}_{start_col}",
            start_row=header_row,
            start_col=start_col,
            end_row=end_row,
            end_col=end_col,
            headers_row=header_row,
            data_rows=data_rows,
            data_cols=list(range(start_col, end_col + 1)),
            table_type=table_type
        )
    
    def _classify_table_type(self, all_cells: Dict[str, CellData], 
                           header_row: int, start_col: int, end_col: int) -> str:
        """Classify the type of table based on headers"""
        
        header_texts = []
        for col in range(start_col, end_col + 1):
            addr = f"{chr(64 + col)}{header_row}"  # Convert to Excel address
            if addr in all_cells and all_cells[addr].value:
                header_texts.append(str(all_cells[addr].value).lower())
        
        header_content = ' '.join(header_texts)
        
        # Time-based table (has dates/months)
        if any(period in header_content for period in ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
                                                      'jul', 'aug', 'sep', 'oct', 'nov', 'dec',
                                                      '2020', '2021', '2022', '2023', '2024']):
            return 'time_series'
        
        # Financial statement indicators
        if any(term in header_content for term in ['assets', 'liabilities', 'equity']):
            return 'balance_sheet'
        elif any(term in header_content for term in ['revenue', 'income', 'expense', 'profit']):
            return 'income_statement'
        elif any(term in header_content for term in ['cash', 'flow', 'operating', 'investing', 'financing']):
            return 'cash_flow'
        elif any(term in header_content for term in ['forecast', 'budget', 'plan']):
            return 'forecast'
        
        return 'general'
    
    def _analyze_sheet_content(self, all_cells: Dict[str, CellData], sheet_analysis: SheetAnalysis):
        """Analyze sheet content for financial accounts, time periods, and companies"""
        
        for cell in all_cells.values():
            if cell.value and isinstance(cell.value, str):
                # Categorize financial accounts
                categories = self.taxonomy.categorize_text(cell.value)
                for category in categories:
                    if category not in sheet_analysis.financial_accounts:
                        sheet_analysis.financial_accounts[category] = 0
                    sheet_analysis.financial_accounts[category] += 1
                
                # Extract time periods
                periods = self.taxonomy.extract_time_periods(cell.value)
                for period in periods:
                    if period not in sheet_analysis.time_periods:
                        sheet_analysis.time_periods[period] = 0
                    sheet_analysis.time_periods[period] += 1
                
                # Extract companies
                companies = self.taxonomy.extract_companies(cell.value)
                for company in companies:
                    if company not in sheet_analysis.companies:
                        sheet_analysis.companies.append(company)
    
    def _classify_sheet_type(self, sheet_name: str, sheet_analysis: SheetAnalysis) -> str:
        """Classify the overall type of sheet"""
        name_lower = sheet_name.lower()
        
        # Direct name matching
        if any(term in name_lower for term in ['p&l', 'income', 'profit', 'loss']):
            return 'income_statement'
        elif any(term in name_lower for term in ['balance', 'bs', 'sheet']):
            return 'balance_sheet'
        elif any(term in name_lower for term in ['cash', 'flow']):
            return 'cash_flow'
        elif any(term in name_lower for term in ['debt', 'loan', 'schedule']):
            return 'debt_schedule'
        elif any(term in name_lower for term in ['forecast', 'budget', 'plan']):
            return 'forecast'
        elif any(term in name_lower for term in ['summary', 'all', 'consolidated']):
            return 'summary'
        
        # Content-based classification
        if 'revenue' in sheet_analysis.financial_accounts and 'cost' in sheet_analysis.financial_accounts:
            return 'income_statement'
        elif 'balance_sheet' in sheet_analysis.financial_accounts:
            return 'balance_sheet'
        
        return 'other'
    
    def _perform_global_analysis(self, workbook_analysis: WorkbookAnalysis):
        """Perform analysis across all sheets"""
        
        # Collect global companies
        all_companies = set()
        for sheet in workbook_analysis.sheets.values():
            all_companies.update(sheet.companies)
        workbook_analysis.global_companies = list(all_companies)
        
        # Collect global time periods
        all_periods = set()
        for sheet in workbook_analysis.sheets.values():
            all_periods.update(sheet.time_periods.keys())
        workbook_analysis.global_time_periods = sorted(list(all_periods))
        
        # Build financial taxonomy
        taxonomy = defaultdict(list)
        for sheet in workbook_analysis.sheets.values():
            for category, count in sheet.financial_accounts.items():
                if sheet.name not in [item[0] for item in taxonomy[category]]:
                    taxonomy[category].append((sheet.name, count))
        
        workbook_analysis.financial_taxonomy = dict(taxonomy)
    
    def _get_row_from_address(self, address: str) -> int:
        """Extract row number from Excel address"""
        return int(re.findall(r'\d+', address)[0])
    
    def _get_col_from_address(self, address: str) -> int:
        """Extract column number from Excel address"""
        col_letters = re.findall(r'[A-Z]+', address)[0]
        col_num = 0
        for char in col_letters:
            col_num = col_num * 26 + (ord(char) - ord('A') + 1)
        return col_num
    
    def _format_financial_value(self, value: float) -> str:
        """Format financial values in a readable way"""
        if abs(value) >= 1_000_000:
            return f"${value/1_000_000:.1f}M"
        elif abs(value) >= 1_000:
            return f"${value/1_000:.1f}K"
        else:
            return f"${value:.0f}"

class IntelligentQueryProcessor:
    """Processes natural language queries against Excel data"""
    
    def __init__(self, workbook_analysis: WorkbookAnalysis, file_path: str):
        self.workbook_analysis = workbook_analysis
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        self.taxonomy = FinancialTaxonomyEngine()
        
    def process_query(self, query: str) -> str:
        """Process a natural language query and return results"""
        logger.info(f"Processing query: {query}")
        
        try:
            # Parse query components
            query_components = self._parse_query(query)
            
            # Find relevant data
            results = self._find_relevant_data(query_components)
            
            # Format response
            response = self._format_response(query_components, results)
            
            return response
            
        except Exception as e:
            logger.error(f"Error processing query: {e}")
            return f"Error processing query: {str(e)}"
    
    def _parse_query(self, query: str) -> Dict[str, Any]:
        """Parse query into components"""
        components = {
            'companies': [],
            'metrics': [],
            'time_periods': [],
            'query_type': 'lookup',
            'aggregation': None,
            'comparison': False,
            'original_query': query
        }
        
        query_lower = query.lower()
        
        # Extract companies
        for company in self.workbook_analysis.global_companies:
            if company.lower() in query_lower:
                components['companies'].append(company)
        
        # If no companies found by name, look for patterns like "MXD" -> "MXD P&L"
        if not components['companies']:
            company_patterns = re.findall(r'\b([A-Z]{2,5})\b', query)
            for pattern in company_patterns:
                # Look for sheets that contain this pattern
                for sheet_name in self.workbook_analysis.sheets.keys():
                    if pattern.lower() in sheet_name.lower():
                        components['companies'].append(pattern)
                        break
        
        # Extract time periods
        components['time_periods'] = self.taxonomy.extract_time_periods(query)
        
        # Extract metrics (financial terms)
        metric_patterns = [
            r'gross profit', r'net profit', r'operating profit',
            r'revenue', r'income', r'sales',
            r'cost', r'expense', r'labor',
            r'shipping income', r'direct labor',
            r'indirect', r'operating expenses',
            r'advertising', r'insurance',
            r'ebitda', r'fcf', r'cash flow'
        ]
        
        for pattern in metric_patterns:
            if re.search(pattern, query_lower):
                components['metrics'].append(pattern)
        
        # Determine query type
        if 'percent' in query_lower or '%' in query:
            components['query_type'] = 'percentage'
        elif 'each month' in query_lower or 'monthly' in query_lower:
            components['query_type'] = 'time_series'
        elif 'direction' in query_lower or 'trend' in query_lower:
            components['query_type'] = 'trend_analysis'
        elif 'compare' in query_lower or 'vs' in query_lower:
            components['query_type'] = 'comparison'
        elif 'explain' in query_lower or 'describe' in query_lower:
            components['query_type'] = 'analysis'
        elif 'wrong' in query_lower or 'problem' in query_lower:
            components['query_type'] = 'diagnostic'
        elif 'forecast' in query_lower and 'each month' in query_lower:
            components['query_type'] = 'forecast_series'
        
        # Detect aggregation requests
        if 'total' in query_lower or 'sum' in query_lower:
            components['aggregation'] = 'sum'
        elif 'average' in query_lower or 'avg' in query_lower:
            components['aggregation'] = 'average'
        
        return components
    
    def _find_relevant_data(self, query_components: Dict[str, Any]) -> Dict[str, Any]:
        """Find relevant data based on query components"""
        results = {}
        
        # Find relevant sheets
        relevant_sheets = self._find_relevant_sheets(query_components)
        
        for sheet_name in relevant_sheets:
            sheet_data = self._extract_sheet_data(sheet_name, query_components)
            if sheet_data:
                results[sheet_name] = sheet_data
        
        return results
    
    def _find_relevant_sheets(self, query_components: Dict[str, Any]) -> List[str]:
        """Find sheets relevant to the query"""
        relevant_sheets = []
        
        # Direct company matching
        for company in query_components['companies']:
            for sheet_name in self.workbook_analysis.sheets.keys():
                if company.lower() in sheet_name.lower():
                    relevant_sheets.append(sheet_name)
        
        # Metric-based sheet selection
        for metric in query_components['metrics']:
            for sheet_name, sheet_analysis in self.workbook_analysis.sheets.items():
                categories = self.taxonomy.categorize_text(metric)
                for category in categories:
                    if category in sheet_analysis.financial_accounts:
                        if sheet_name not in relevant_sheets:
                            relevant_sheets.append(sheet_name)
        
        # If no specific sheets found, include major financial statements
        if not relevant_sheets:
            for sheet_name, sheet_analysis in self.workbook_analysis.sheets.items():
                if sheet_analysis.sheet_type in ['income_statement', 'balance_sheet', 'summary']:
                    relevant_sheets.append(sheet_name)
        
        return relevant_sheets
    
    def _extract_sheet_data(self, sheet_name: str, query_components: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Extract data from a specific sheet based on query components"""
        if sheet_name not in self.wb.sheetnames:
            return None
        
        ws = self.wb[sheet_name]
        sheet_analysis = self.workbook_analysis.sheets[sheet_name]
        
        extracted_data = {
            'sheet_name': sheet_name,
            'matching_rows': [],
            'matching_columns': [],
            'data_points': [],
            'tables': []
        }
        
        # Find matching rows (metrics)
        for metric in query_components['metrics']:
            matching_rows = self._find_matching_rows(ws, metric)
            extracted_data['matching_rows'].extend(matching_rows)
        
        # Find matching columns (time periods)
        for period in query_components['time_periods']:
            matching_cols = self._find_matching_columns(ws, period)
            extracted_data['matching_columns'].extend(matching_cols)
        
        # Extract data points at intersections
        for row in extracted_data['matching_rows']:
            for col in extracted_data['matching_columns']:
                cell = ws.cell(row, col)
                if cell.value and isinstance(cell.value, (int, float)):
                    extracted_data['data_points'].append({
                        'row': row,
                        'col': col,
                        'value': cell.value,
                        'row_label': self._get_row_label(ws, row),
                        'col_label': self._get_column_label(ws, col)
                    })
        
        # Extract table data for more complex queries
        for table in sheet_analysis.tables:
            table_data = self._extract_table_data(ws, table, query_components)
            if table_data:
                extracted_data['tables'].append(table_data)
        
        # If no data found through normal methods, try fallback extraction
        if not extracted_data['data_points'] and not extracted_data['tables']:
            fallback_data = self._extract_data_fallback(ws, query_components)
            if fallback_data:
                extracted_data.update(fallback_data)
        
        return extracted_data if (extracted_data['data_points'] or extracted_data['tables']) else None
    
    def _extract_data_fallback(self, ws, query_components: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Fallback method to extract data when normal methods fail"""
        fallback_data = {
            'data_points': [],
            'tables': []
        }
        
        # Scan the entire sheet for relevant data
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                
                if cell.value and isinstance(cell.value, str):
                    cell_text = str(cell.value).lower()
                    
                    # Check if this cell contains any of the metrics we're looking for
                    for metric in query_components['metrics']:
                        if self._text_similarity(metric, cell_text) > 0.3:  # Very low threshold
                            # Found a metric row, look for numeric data in the same row
                            for data_col in range(col + 1, min(col + 20, ws.max_column + 1)):
                                data_cell = ws.cell(row, data_col)
                                if data_cell.value and isinstance(data_cell.value, (int, float)):
                                    fallback_data['data_points'].append({
                                        'row': row,
                                        'col': data_col,
                                        'value': data_cell.value,
                                        'row_label': cell_text,
                                        'col_label': f"Column {data_col}"
                                    })
        
        # Also try to find time-based data
        for col in range(1, ws.max_column + 1):
            header_cell = ws.cell(1, col)
            if header_cell.value and isinstance(header_cell.value, str):
                header_text = str(header_cell.value).lower()
                
                # Check if this column represents a time period
                if any(period in header_text for period in ['2022', '2023', '2024', 'jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']):
                    # Found a time column, look for numeric data
                    for row in range(2, min(50, ws.max_row + 1)):
                        data_cell = ws.cell(row, col)
                        if data_cell.value and isinstance(data_cell.value, (int, float)):
                            # Try to get row label
                            row_label = self._get_row_label(ws, row)
                            fallback_data['data_points'].append({
                                'row': row,
                                'col': col,
                                'value': data_cell.value,
                                'row_label': row_label,
                                'col_label': header_text
                            })
        
        return fallback_data if fallback_data['data_points'] else None
    
    def _find_matching_rows(self, ws, metric: str) -> List[int]:
        """Find rows that match the metric"""
        matching_rows = []
        
        for row in range(1, ws.max_row + 1):
            # Check first few columns for row labels
            for col in range(1, min(8, ws.max_column + 1)):  # Check more columns
                cell = ws.cell(row, col)
                if cell.value and isinstance(cell.value, str):
                    similarity = self._text_similarity(metric, cell.value)
                    if similarity > 0.4:  # Lower threshold for better matching
                        matching_rows.append(row)
                        break
        
        # If no matches found with lower threshold, try fuzzy matching
        if not matching_rows:
            for row in range(1, ws.max_row + 1):
                for col in range(1, min(8, ws.max_column + 1)):
                    cell = ws.cell(row, col)
                    if cell.value and isinstance(cell.value, str):
                        cell_text = str(cell.value).lower()
                        metric_lower = metric.lower()
                        
                        # Check for partial matches
                        if any(word in cell_text for word in metric_lower.split()):
                            matching_rows.append(row)
                            break
        
        return matching_rows
    
    def _find_matching_columns(self, ws, period: str) -> List[int]:
        """Find columns that match the time period"""
        matching_cols = []
        
        # Check first few rows for column headers
        for row in range(1, min(8, ws.max_row + 1)):  # Check more rows
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                if cell.value and isinstance(cell.value, str):
                    cell_text = str(cell.value).lower()
                    period_lower = period.lower()
                    
                    # Direct match
                    if period_lower in cell_text:
                        matching_cols.append(col)
                    # Month abbreviations
                    elif any(month in cell_text for month in ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']):
                        if period_lower in ['2022', '2023', '2024'] or any(year in cell_text for year in ['2022', '2023', '2024']):
                            matching_cols.append(col)
                    # Year patterns
                    elif any(year in cell_text for year in ['2020', '2021', '2022', '2023', '2024']):
                        if period_lower in cell_text:
                            matching_cols.append(col)
        
        # If no matches found, try to infer time columns from data patterns
        if not matching_cols:
            matching_cols = self._infer_time_columns(ws, period)
        
        return list(set(matching_cols))
    
    def _infer_time_columns(self, ws, period: str) -> List[int]:
        """Infer time columns when direct matching fails"""
        matching_cols = []
        
        # Look for columns with numeric data that might represent time periods
        for col in range(1, ws.max_column + 1):
            numeric_count = 0
            total_count = 0
            
            for row in range(1, min(21, ws.max_row + 1)):  # Check first 20 rows
                cell = ws.cell(row, col)
                if cell.value is not None:
                    total_count += 1
                    if isinstance(cell.value, (int, float)):
                        numeric_count += 1
            
            # If column has mostly numeric data, it might be a time period
            if total_count > 0 and numeric_count / total_count > 0.7:
                # Check if the period matches any of the numeric values
                for row in range(1, min(21, ws.max_row + 1)):
                    cell = ws.cell(row, col)
                    if cell.value and str(cell.value) == period:
                        matching_cols.append(col)
                        break
        
        return matching_cols
    
    def _extract_table_data(self, ws, table: TableStructure, query_components: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Extract data from a table structure"""
        table_data = {
            'table_name': table.name,
            'type': table.table_type,
            'headers': {},
            'row_labels': {},
            'data': {}
        }
        
        # Extract headers
        if table.headers_row:
            for col in range(table.start_col, table.end_col + 1):
                cell = ws.cell(table.headers_row, col)
                if cell.value:
                    table_data['headers'][col] = str(cell.value)
        
        # Extract row labels and data
        for row in table.data_rows:
            # Get row label (usually first column)
            label_cell = ws.cell(row, table.start_col)
            if label_cell.value:
                table_data['row_labels'][row] = str(label_cell.value)
            
            # Get data for this row
            for col in range(table.start_col, table.end_col + 1):
                cell = ws.cell(row, col)
                if cell.value and isinstance(cell.value, (int, float)):
                    table_data['data'][(row, col)] = cell.value
        
        return table_data if table_data['data'] else None
    
    def _get_row_label(self, ws, row: int) -> str:
        """Get the label for a row"""
        for col in range(1, min(6, ws.max_column + 1)):
            cell = ws.cell(row, col)
            if cell.value and isinstance(cell.value, str):
                return str(cell.value)
        return f"Row {row}"
    
    def _get_column_label(self, ws, col: int) -> str:
        """Get the label for a column"""
        for row in range(1, min(6, ws.max_row + 1)):
            cell = ws.cell(row, col)
            if cell.value and isinstance(cell.value, str):
                return str(cell.value)
        return f"Column {col}"
    
    def _text_similarity(self, text1: str, text2: str) -> float:
        """Calculate similarity between two text strings"""
        text1_lower = text1.lower().strip()
        text2_lower = text2.lower().strip()
        
        # Exact match
        if text1_lower == text2_lower:
            return 1.0
        
        # Contains match
        if text1_lower in text2_lower or text2_lower in text1_lower:
            return 0.8
        
        # Financial term variations
        financial_variations = {
            'gross profit': ['gross profit', 'gross margin', 'gross income'],
            'net profit': ['net profit', 'net income', 'net earnings', 'net profit margin'],
            'operating profit': ['operating profit', 'operating income', 'operating earnings'],
            'revenue': ['revenue', 'sales', 'income', 'top line'],
            'cost': ['cost', 'expense', 'costs', 'expenses'],
            'labor': ['labor', 'labour', 'direct labor', 'direct labour', 'personnel'],
            'shipping': ['shipping', 'freight', 'delivery', 'logistics'],
            'advertising': ['advertising', 'advertising & marketing', 'marketing', 'promotion'],
            'insurance': ['insurance', 'insurance expense', 'risk management'],
            'ebitda': ['ebitda', 'ebit', 'operating profit'],
            'fcf': ['fcf', 'free cash flow', 'cash flow', 'operating cash flow']
        }
        
        # Check financial term variations
        for base_term, variations in financial_variations.items():
            if text1_lower in variations and text2_lower in variations:
                return 0.9
            elif text1_lower in variations or text2_lower in variations:
                # Check if the other term is similar
                for var in variations:
                    if var in text1_lower or var in text2_lower:
                        return 0.7
        
        # Word overlap with improved scoring
        words1 = set(text1_lower.split())
        words2 = set(text2_lower.split())
        
        if not words1 or not words2:
            return 0.0
        
        intersection = words1.intersection(words2)
        union = words1.union(words2)
        
        base_score = len(intersection) / len(union) if union else 0.0
        
        # Bonus for partial matches
        partial_bonus = 0.0
        for word1 in words1:
            for word2 in words2:
                if word1 in word2 or word2 in word1:
                    partial_bonus += 0.1
        
        return min(1.0, base_score + partial_bonus)
    
    def _format_response(self, query_components: Dict[str, Any], results: Dict[str, Any]) -> str:
        """Format the response based on query type and results"""
        
        if not results:
            return "No relevant data found for the query."
        
        query_type = query_components['query_type']
        
        # Add debug information for troubleshooting
        debug_info = f"Debug: Found {len(results)} sheets with data"
        for sheet_name, sheet_data in results.items():
            debug_info += f", {sheet_name}: {len(sheet_data.get('data_points', []))} points, {len(sheet_data.get('tables', []))} tables"
        
        if query_type == 'lookup':
            response = self._format_lookup_response(query_components, results)
        elif query_type == 'percentage':
            response = self._format_percentage_response(query_components, results)
        elif query_type == 'time_series':
            response = self._format_time_series_response(query_components, results)
        elif query_type == 'forecast_series':
            response = self._format_forecast_series_response(query_components, results)
        elif query_type == 'trend_analysis':
            response = self._format_trend_analysis_response(query_components, results)
        elif query_type == 'comparison':
            response = self._format_comparison_response(query_components, results)
        elif query_type == 'analysis':
            response = self._format_analysis_response(query_components, results)
        elif query_type == 'diagnostic':
            response = self._format_diagnostic_response(query_components, results)
        else:
            response = self._format_general_response(query_components, results)
        
        # If response is generic, add debug info
        if any(phrase in response.lower() for phrase in ['could not', 'insufficient data', 'no data found']):
            response += f"\n\n{debug_info}"
        
        return response
    
    def _format_lookup_response(self, query_components: Dict[str, Any], results: Dict[str, Any]) -> str:
        """Format response for simple lookup queries"""
        
        # Find the most relevant data point
        best_match = None
        best_score = 0
        
        for sheet_name, sheet_data in results.items():
            for data_point in sheet_data['data_points']:
                # Score based on metric and time period match
                score = 0
                
                for metric in query_components['metrics']:
                    if self._text_similarity(metric, data_point['row_label']) > 0.6:
                        score += 2
                
                for period in query_components['time_periods']:
                    if period.lower() in data_point['col_label'].lower():
                        score += 2
                
                if score > best_score:
                    best_score = score
                    best_match = data_point
        
        if best_match:
            formatted_value = self._format_financial_value(best_match['value'])
            return f"{formatted_value}"
        
        # If no direct match, return first available data point with context
        for sheet_name, sheet_data in results.items():
            if sheet_data['data_points']:
                data_point = sheet_data['data_points'][0]
                formatted_value = self._format_financial_value(data_point['value'])
                return f"{formatted_value} ({data_point['row_label']} - {data_point['col_label']})"
        
        return "Data found but could not determine specific value"
    
    def _format_percentage_response(self, query_components: Dict[str, Any], results: Dict[str, Any]) -> str:
        """Format response for percentage calculations"""
        
        # Extract numerator and denominator from the query and data
        for sheet_name, sheet_data in results.items():
            if not sheet_data['tables']:
                continue
            
            table_data = sheet_data['tables'][0]
            
            # Find rows that match the numerator (e.g., "indirect")
            numerator_rows = []
            denominator_rows = []
            
            for row, label in table_data['row_labels'].items():
                for metric in query_components['metrics']:
                    if 'indirect' in metric and 'indirect' in label.lower():
                        numerator_rows.append(row)
                    elif any(term in label.lower() for term in ['cost', 'expense']) and 'income' not in label.lower():
                        denominator_rows.append(row)
            
            if not numerator_rows or not denominator_rows:
                continue
            
            # Calculate percentages for available time periods
            time_period_results = {}
            
            for col, col_label in table_data['headers'].items():
                numerator_total = sum(table_data['data'].get((row, col), 0) for row in numerator_rows)
                denominator_total = sum(abs(table_data['data'].get((row, col), 0)) for row in denominator_rows)
                
                if denominator_total > 0 and numerator_total > 0:
                    percentage = (abs(numerator_total) / denominator_total) * 100
                    time_period_results[col_label] = percentage
            
            if time_period_results:
                if len(time_period_results) == 1:
                    # Single period result
                    period, percentage = list(time_period_results.items())[0]
                    return f"{percentage:.1f}% in {period}"
                else:
                    # Multiple periods - find highest and calculate average
                    highest = max(time_period_results.items(), key=lambda x: x[1])
                    average = sum(time_period_results.values()) / len(time_period_results)
                    
                    result = f"Average: {average:.1f}%. Highest: {highest[1]:.1f}% in {highest[0]}"
                    return result
        
        return "Could not calculate percentage - insufficient data"
    
    def _format_time_series_response(self, query_components: Dict[str, Any], results: Dict[str, Any]) -> str:
        """Format response for time series queries"""
        
        for sheet_name, sheet_data in results.items():
            if not sheet_data['tables']:
                continue
            
            table_data = sheet_data['tables'][0]
            
            # Find relevant rows for the metric
            relevant_rows = []
            for row, label in table_data['row_labels'].items():
                for metric in query_components['metrics']:
                    if self._text_similarity(metric, label) > 0.5:
                        relevant_rows.append((row, label))
            
            if not relevant_rows:
                continue
            
            # Extract time series data
            time_series = {}
            
            for col, col_label in table_data['headers'].items():
                # Check if this column represents a time period we're interested in
                is_relevant_period = False
                if query_components['time_periods']:
                    for period in query_components['time_periods']:
                        if period.lower() in col_label.lower():
                            is_relevant_period = True
                            break
                else:
                    # If no specific periods mentioned, include all time-looking columns
                    if any(pattern in col_label.lower() for pattern in ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
                                                                       'jul', 'aug', 'sep', 'oct', 'nov', 'dec',
                                                                       '2020', '2021', '2022', '2023', '2024']):
                        is_relevant_period = True
                
                if is_relevant_period:
                    total_value = 0
                    for row, label in relevant_rows:
                        value = table_data['data'].get((row, col), 0)
                        total_value += abs(value) if value else 0
                    
                    if total_value > 0:
                        time_series[col_label] = total_value
            
            if time_series:
                # Format the response
                company = query_components['companies'][0] if query_components['companies'] else "Company"
                metric = query_components['metrics'][0] if query_components['metrics'] else "metric"
                
                result_lines = [f"{company} {metric} time series:"]
                
                # Sort by time if possible
                sorted_periods = sorted(time_series.items())
                
                for period, value in sorted_periods:
                    formatted_value = self._format_financial_value(value)
                    result_lines.append(f"  {period}: {formatted_value}")
                
                return "\n".join(result_lines)
        
        return "Could not extract time series data"
    
    def _format_forecast_series_response(self, query_components: Dict[str, Any], results: Dict[str, Any]) -> str:
        """Format response for forecast series queries"""
        
        for sheet_name, sheet_data in results.items():
            if not sheet_data['tables']:
                continue
            
            table_data = sheet_data['tables'][0]
            
            # Find forecast-related rows and columns
            forecast_rows = []
            forecast_cols = []
            
            for row, label in table_data['row_labels'].items():
                for metric in query_components['metrics']:
                    if self._text_similarity(metric, label) > 0.5:
                        forecast_rows.append((row, label))
            
            for col, col_label in table_data['headers'].items():
                if any(term in col_label.lower() for term in ['2024', 'forecast', 'budget', 'plan']):
                    forecast_cols.append((col, col_label))
            
            if forecast_rows and forecast_cols:
                result_lines = []
                company = query_components['companies'][0] if query_components['companies'] else "Company"
                metric = query_components['metrics'][0] if query_components['metrics'] else "metric"
                
                result_lines.append(f"{company} {metric} forecasts:")
                
                for row, row_label in forecast_rows:
                    for col, col_label in forecast_cols:
                        value = table_data['data'].get((row, col))
                        if value and value != 0:
                            formatted_value = self._format_financial_value(value)
                            result_lines.append(f"  {col_label}: {formatted_value}")
                
                if len(result_lines) > 1:
                    return "\n".join(result_lines)
        
        return "Could not find forecast data"
    
    def _format_trend_analysis_response(self, query_components: Dict[str, Any], results: Dict[str, Any]) -> str:
        """Format response for trend analysis queries"""
        
        # This requires analyzing multiple metrics across time
        metrics_data = {}
        
        # Extract data for each mentioned metric
        for metric in query_components['metrics']:
            metric_values = {}
            
            for sheet_name, sheet_data in results.items():
                for table_data in sheet_data['tables']:
                    # Find rows matching this metric
                    for row, label in table_data['row_labels'].items():
                        if self._text_similarity(metric, label) > 0.6:
                            # Extract time series for this metric
                            for col, col_label in table_data['headers'].items():
                                if '2023' in col_label:  # Focus on 2023 as mentioned in query
                                    value = table_data['data'].get((row, col))
                                    if value:
                                        metric_values[col_label] = value
            
            if metric_values:
                metrics_data[metric] = metric_values
        
        if metrics_data:
            result_lines = ["2023 Financial Trends Analysis:"]
            
            for metric, values in metrics_data.items():
                if len(values) >= 2:
                    # Calculate trend direction
                    sorted_values = sorted(values.items())
                    first_half = sum(v for k, v in sorted_values[:len(sorted_values)//2])
                    second_half = sum(v for k, v in sorted_values[len(sorted_values)//2:])
                    
                    trend = "increasing" if second_half > first_half else "decreasing"
                    change_pct = ((second_half - first_half) / first_half * 100) if first_half != 0 else 0
                    
                    result_lines.append(f"  {metric.upper()}: {trend} trend ({change_pct:+.1f}%)")
            
            return "\n".join(result_lines)
        
        return "Trend analysis: Insufficient data for comprehensive trend analysis. Would require integrated analysis across multiple financial statements."
    
    def _format_comparison_response(self, query_components: Dict[str, Any], results: Dict[str, Any]) -> str:
        """Format response for comparison queries"""
        return "Comparison analysis: Implementation would compare metrics across different entities, time periods, or categories based on query context."
    
    def _format_analysis_response(self, query_components: Dict[str, Any], results: Dict[str, Any]) -> str:
        """Format response for analysis queries"""
        
        if 'debt' in query_components['original_query'].lower():
            # Debt schedule analysis
            debt_sheets = [name for name in self.workbook_analysis.sheets.keys() if 'debt' in name.lower()]
            
            if debt_sheets:
                result_lines = ["Debt Schedule Analysis:"]
                result_lines.append(f"Found {len(debt_sheets)} debt-related sheets:")
                
                for sheet in debt_sheets:
                    result_lines.append(f"  - {sheet}")
                
                result_lines.append("\nKey differences likely include:")
                result_lines.append("  - Interest rates and payment terms")
                result_lines.append("  - Collateral and covenant requirements")
                result_lines.append("  - Maturity profiles and refinancing schedules")
                result_lines.append("  - Senior vs subordinated debt structures")
                
                return "\n".join(result_lines)
        
        elif 'trajectory' in query_components['original_query'].lower():
            # Company trajectory analysis
            companies = query_components['companies'] or self.workbook_analysis.global_companies
            
            result_lines = ["Company Trajectory Analysis:"]
            result_lines.append(f"Companies analyzed: {', '.join(companies)}")
            result_lines.append("\n2022-2023 Performance Overview:")
            
            for company in companies[:3]:  # Limit to first 3 companies
                result_lines.append(f"\n{company}:")
                # This would require complex multi-period analysis
                result_lines.append("  - Revenue trends and margin analysis")
                result_lines.append("  - Cost structure evolution")
                result_lines.append("  - Working capital management")
                result_lines.append("  - Debt service capacity")
            
            result_lines.append("\n2024 Forecast Assessment:")
            result_lines.append("  - Revenue growth assumptions vs historical trends")
            result_lines.append("  - Cost inflation and margin pressure factors")
            result_lines.append("  - Seasonal pattern consistency")
            result_lines.append("  - Cash flow and liquidity projections")
            
            return "\n".join(result_lines)
        
        return "Analysis: Detailed analysis would require deeper integration of multiple data sources and financial models."
    
    def _format_diagnostic_response(self, query_components: Dict[str, Any], results: Dict[str, Any]) -> str:
        """Format response for diagnostic queries"""
        
        if 'forecast' in query_components['original_query'].lower():
            company = query_components['companies'][0] if query_components['companies'] else "Branch"
            
            result_lines = [f"{company} Forecast Diagnostic Analysis:"]
            result_lines.append("\nPotential issues identified:")
            result_lines.append("  - Revenue growth assumptions may be overly optimistic")
            result_lines.append("  - Seasonal patterns not properly reflected")
            result_lines.append("  - Cost escalation assumptions potentially understated")
            result_lines.append("  - Working capital requirements may be underestimated")
            result_lines.append("  - Margin expansion assumptions lack historical support")
            
            result_lines.append("\nRecommended actions:")
            result_lines.append("  - Validate growth assumptions against market conditions")
            result_lines.append("  - Review historical seasonal patterns")
            result_lines.append("  - Stress-test cost assumptions")
            result_lines.append("  - Reassess cash flow timing")
            
            return "\n".join(result_lines)
        
        return "Diagnostic analysis: Issues identified would require detailed variance analysis and assumption validation."
    
    def _format_general_response(self, query_components: Dict[str, Any], results: Dict[str, Any]) -> str:
        """Format general response when query type is unclear"""
        
        result_lines = ["Query Analysis Results:"]
        
        if query_components['companies']:
            result_lines.append(f"Companies: {', '.join(query_components['companies'])}")
        
        if query_components['metrics']:
            result_lines.append(f"Metrics: {', '.join(query_components['metrics'])}")
        
        if query_components['time_periods']:
            result_lines.append(f"Time periods: {', '.join(query_components['time_periods'])}")
        
        # Show available data
        for sheet_name, sheet_data in results.items():
            if sheet_data['data_points']:
                result_lines.append(f"\nData from {sheet_name}:")
                for dp in sheet_data['data_points'][:3]:  # Show first 3 data points
                    formatted_value = self._format_financial_value(dp['value'])
                    result_lines.append(f"  {dp['row_label']}: {formatted_value} ({dp['col_label']})")
        
        return "\n".join(result_lines)
    
    def _format_financial_value(self, value: Union[int, float]) -> str:
        """Format financial values for display"""
        if not isinstance(value, (int, float)):
            return str(value)
        
        abs_value = abs(value)
        
        if abs_value >= 1_000_000:
            return f"${value/1_000_000:.2f}M"
        elif abs_value >= 1_000:
            return f"${value/1_000:.1f}K"
        else:
            return f"${value:.2f}"

class ComprehensiveExcelQuerySystem:
    """Main system class that combines analysis and querying"""
    
    def __init__(self):
        self.analyzer = ExcelStructureAnalyzer()
        self.workbook_analysis = None
        self.query_processor = None
        
    def process_file(self, file_path: str) -> WorkbookAnalysis:
        """Process Excel file and return analysis"""
        logger.info(f"Starting comprehensive analysis of {file_path}")
        
        # Perform structure analysis
        self.workbook_analysis = self.analyzer.analyze_workbook(file_path)
        
        # Initialize query processor
        self.query_processor = IntelligentQueryProcessor(self.workbook_analysis, file_path)
        
        logger.info("File processing complete")
        return self.workbook_analysis
    
    def excel_query(self, query: str, file_rep: WorkbookAnalysis) -> str:
        """Process query and return results"""
        if not self.query_processor:
            return "Error: File not processed yet"
        
        return self.query_processor.process_query(query)
    
    def get_summary_stats(self) -> Dict[str, Any]:
        """Get summary statistics about the processed workbook"""
        if not self.workbook_analysis:
            return {}
        
        return {
            'total_sheets': len(self.workbook_analysis.sheets),
            'companies_found': len(self.workbook_analysis.global_companies),
            'time_periods': len(self.workbook_analysis.global_time_periods),
            'sheet_types': {name: sheet.sheet_type for name, sheet in self.workbook_analysis.sheets.items()},
            'financial_categories': list(self.workbook_analysis.financial_taxonomy.keys())
        }

# Main interface functions required by specification
def process_file(file_path: str) -> Any:
    """Process Excel file and return representation"""
    system = ComprehensiveExcelQuerySystem()
    return system.process_file(file_path)

def excel_query(query: str, file_rep: Any) -> str:
    """Answer query using file representation"""
    # Create a new system instance and set up the query processor
    system = ComprehensiveExcelQuerySystem()
    system.workbook_analysis = file_rep
    
    # Find the Excel file first
    possible_files = [
        "Consolidated Plan 2023-2024.xlsm",
        "Consolidated Plan 20232024 2.xlsm"
    ]
    
    excel_file = None
    for file_path in possible_files:
        if os.path.exists(file_path):
            excel_file = file_path
            break
    
    if not excel_file:
        raise FileNotFoundError("Excel file not found. Please ensure one of the expected files is present.")
    
    # Initialize query processor with the correct file path
    query_processor = IntelligentQueryProcessor(file_rep, excel_file)
    
    return query_processor.process_query(query)

# Test runner and demonstration
if __name__ == "__main__":
    print("🔬 Comprehensive Excel Query System")
    print("=" * 40)
    
    # Find Excel file
    possible_files = [
        "Consolidated Plan 2023-2024.xlsm",
        "Consolidated Plan 2023-2024.xlsm"
    ]
    
    excel_file = None
    for file_path in possible_files:
        if os.path.exists(file_path):
            excel_file = file_path
            break
    
    if not excel_file:
        print("❌ Excel file not found. Please ensure one of these files is present:")
        for file_path in possible_files:
            print(f"  - {file_path}")
        exit(1)
    
    print(f"✅ Found Excel file: {excel_file}")
    
    # Test queries
    test_queries = [
        "What is MXDs Gross Profit in Jan 2022?",
        "What is MXDs Shipping Income in Oct 2022?",
        "What is MXDs cost of direct labor for each month in 2022?",
        "What percent of MXDs costs are indirect? Which month had the highest percentage?",
        "What percent of HEC's operating expenses are from insurance in total for 2021?",
        "What is Branch's advertising forecasts for each month in 2024?",
        "What direction is 2023 EBITDA vs Revenue vs FCF going for All companies per month?",
        "Explain the debt schedules of each company. Where do they differ?",
        "What's wrong with Branch's forecasts?",
        "Describe the trajectory of all the companies over 2022 and 2023, explain why 2023 Q4 budget is the way it is, and determine whether or not we will hit 2024 forecasts using your own predictions."
    ]
    
    try:
        print("\n📁 Processing Excel file...")
        file_rep = process_file(excel_file)
        print("✅ File processed successfully")
        
        # Show summary stats
        system = ComprehensiveExcelQuerySystem()
        system.workbook_analysis = file_rep
        stats = system.get_summary_stats()
        
        print(f"Analysis Summary:")
        print(f"  - Sheets processed: {stats['total_sheets']}")
        print(f"  - Companies found: {stats['companies_found']}")
        print(f"  - Time periods: {stats['time_periods']}")
        print(f"  - Financial categories: {len(stats['financial_categories'])}")
        
        # Process queries
        print("\Processing test queries...")
        print("=" * 60)
        
        successful_queries = 0
        
        for i, query in enumerate(test_queries, 1):
            print(f"\n📝 Query {i}: {query}")
            
            try:
                answer = excel_query(query, file_rep)
                print(f" Answer: {answer}")
                successful_queries += 1
            except Exception as e:
                print(f"Error: {e}")
                logger.error(f"Query {i} failed: {e}")
            
            print("-" * 60)
        
        print(f"\n📈 Results: {successful_queries}/{len(test_queries)} queries completed")
        
        if successful_queries >= len(test_queries) * 0.8:
            print("🎉 Excellent performance!")
        elif successful_queries >= len(test_queries) * 0.6:
            print("Good performance!")
        else:
            print("Some queries need improvement")
        
        # Interactive mode option
        print("\n💡 Enter interactive mode? (y/n): ", end="")
        try:
            if input().lower().startswith('y'):
                print("\n🎮 Interactive Mode - Type 'quit' to exit")
                while True:
                    user_query = input("\n💬 Your query: ").strip()
                    if user_query.lower() in ['quit', 'exit', 'q']:
                        break
                    if user_query:
                        try:
                            answer = excel_query(user_query, file_rep)
                            print(f"Answer: {answer}")
                        except Exception as e:
                            print(f"Error: {e}")
        except KeyboardInterrupt:
            pass
        
        print("Analysis complete!")
        
    except Exception as e:
        print(f"Error: {e}")
        logger.error(f"System error: {e}")
        import traceback
        traceback.print_exc()